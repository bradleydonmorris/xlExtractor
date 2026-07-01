from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ._types import RowParser, ParseResult, Accumulator, ParserType
from .parsers import icase_compare_list, identify_parser


def parse_worksheet(
	file_path: Path,
	sheet_name: str,
	row_parsers: list[RowParser],
) -> list[ParseResult]:
	results: list[ParseResult] = []
	work_book: Workbook = openpyxl.load_workbook(file_path, data_only=True)
	sheet_found: bool = False

	for work_sheet_name in work_book.sheetnames:
		if work_sheet_name.casefold() != sheet_name.casefold():
			continue
		sheet_found = True
		work_sheet: Worksheet = work_book[work_sheet_name]
		accumulator: Accumulator | None = None

		for row_index, row in enumerate(work_sheet.rows):
			if (
				accumulator
				and row_index > accumulator.start_index
				and accumulator.parser.parser_type == ParserType.LOOP_UNTIL_BROKEN
				and accumulator.parser.loop_breakers
				and icase_compare_list(row[0], accumulator.parser.loop_breakers)
			):
				accumulator.force_complete()

			if accumulator is None:
				if parser := identify_parser(row_parsers, row):
					accumulator = Accumulator(parser=parser, start_index=row_index)
					if parser.parser_type != ParserType.LOOP_UNTIL_BROKEN:
						accumulator.add(row)
			else:
				if not accumulator.is_complete:
					accumulator.add(row)

			if accumulator and accumulator.is_complete:
				results.append(accumulator.finalize())
				accumulator = None

		break  # target sheet found; no need to iterate further

	work_book.close()

	if not sheet_found:
		raise ValueError(f"Sheet not found: {sheet_name!r}")
	return results


def get_worksheet_name(file_path: Path, sheet_name: str | None) -> str | None:
	work_book: Workbook = openpyxl.load_workbook(file_path, data_only=True)
	result: str | None = None
	if not sheet_name:
		result = work_book.sheetnames[0]
	else:
		for work_sheet_name in work_book.sheetnames:
			if work_sheet_name.casefold() != sheet_name.casefold():
				continue
			result = work_sheet_name
			break
	work_book.close()
	return result


def report_data_types(results: list[ParseResult]) -> None:
	data_key_types: dict[tuple[str, str], list[type]] = defaultdict(list)
	for parsed_result in results:
		data = parsed_result["data"]
		records: list[dict[str, Any]] = [data] if isinstance(data, dict) else data if isinstance(data, list) else []
		for record in records:
			for key, value in record.items():
				data_key: tuple[str, str] = (parsed_result["row_type"], key)
				t = type(value)
				if t not in data_key_types[data_key]:
					data_key_types[data_key].append(t)
	print(f"{'Row Type'.ljust(25)} {'Field Name'.ljust(25)} Types")
	print(f"{'-'*25} {'-'*25} {'-'*25}")
	for (row_type, field_name), types in data_key_types.items():
		types_str: str = " | ".join(t.__name__ for t in types)
		print(f"{str(row_type).ljust(25)} {field_name.ljust(25)} {types_str}")
